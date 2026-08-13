"""Export van de prognose (incl. scenario's) naar CSV of Excel.

De dagregels komen uit de prognose op de server; de scenariopaden worden door
de frontend aangeleverd, omdat de bijstellingsberekening daar leeft — die moet
tijdens het slepen direct reageren. Alles wat binnenkomt wordt hier gevalideerd,
zodat een kapotte client geen onzin in het bestand kan zetten.
"""

from __future__ import annotations

import io
import math
import re
from datetime import date

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

MAX_SCENARIOS = 10
MAX_NAME_LEN = 60
# Ruime bovengrens; vangt alleen absurde waarden af (oneindig, overflow).
MAX_VALUE = 1e15

DAY_NAMES = ["maandag", "dinsdag", "woensdag", "donderdag", "vrijdag", "zaterdag", "zondag"]

BASE_COLUMNS = [
    "Datum", "Weekdag", "Soort", "Bijzonderheid",
    "Werkelijk", "Prognose", "Ondergrens (10%)", "Bovengrens (90%)",
    "Realisatie vorig jaar", "Verwacht (basis)",
]


class ExportError(ValueError):
    """Ongeldige exportaanvraag."""


def _clean_number(value, label: str) -> float:
    try:
        num = float(value)
    except (TypeError, ValueError):
        raise ExportError(f"Ongeldige waarde in {label}.")
    if not math.isfinite(num):
        raise ExportError(f"Niet-eindige waarde in {label}.")
    if num < 0 or num > MAX_VALUE:
        raise ExportError(f"Waarde buiten bereik in {label}.")
    return num


def clean_scenarios(raw, valid_dates: set[str]) -> list[dict]:
    """Valideert de aangeleverde scenario's en normaliseert ze."""
    if raw is None:
        return []
    if not isinstance(raw, list):
        raise ExportError("Scenario's moeten een lijst zijn.")
    if len(raw) > MAX_SCENARIOS:
        raise ExportError(f"Maximaal {MAX_SCENARIOS} scenario's per export.")

    out, seen = [], set()
    for i, item in enumerate(raw):
        if not isinstance(item, dict):
            raise ExportError("Elk scenario moet een object zijn.")
        name = str(item.get("name") or f"Scenario {i + 1}").strip()[:MAX_NAME_LEN]
        # Kolomnamen moeten uniek zijn, anders overschrijven ze elkaar.
        base, n = name, 2
        while name.lower() in seen:
            name = f"{base} ({n})"
            n += 1
        seen.add(name.lower())

        path = item.get("path") or {}
        if not isinstance(path, dict):
            raise ExportError(f"Scenario '{name}' heeft een ongeldig pad.")
        clean_path = {}
        for iso, value in path.items():
            if iso not in valid_dates:
                continue  # datum buiten de horizon: negeren, niet crashen
            clean_path[iso] = _clean_number(value, f"scenario '{name}'")
        out.append({"name": name, "path": clean_path})
    return out


def build_table(forecast: dict, scenarios: list[dict]) -> pd.DataFrame:
    """Bouwt de exporttabel: dagregels plus één kolom per scenario."""
    rows = []
    for day in forecast["days"]:
        iso = day["date"]
        d = date.fromisoformat(iso)
        actual = day.get("actual")
        pred = day.get("pred")
        bijzonder = day.get("holiday") or ("weekend" if day.get("weekend") else "")
        row = {
            "Datum": iso,
            "Weekdag": DAY_NAMES[d.weekday()],
            "Soort": "realisatie" if actual is not None else "prognose",
            "Bijzonderheid": bijzonder,
            "Werkelijk": actual,
            "Prognose": pred,
            "Ondergrens (10%)": day.get("lo"),
            "Bovengrens (90%)": day.get("hi"),
            "Realisatie vorig jaar": day.get("lastYearActual"),
            # Wat je verwacht zonder bijstellingen: realisatie als die er is,
            # anders de modelprognose.
            "Verwacht (basis)": actual if actual is not None else pred,
        }
        for sc in scenarios:
            # Een scenario verandert niets aan al gerealiseerde dagen.
            row[sc["name"]] = actual if actual is not None else sc["path"].get(iso, pred)
        rows.append(row)

    columns = BASE_COLUMNS + [sc["name"] for sc in scenarios]
    return pd.DataFrame(rows, columns=columns)


def _totals(table: pd.DataFrame, forecast: dict, scenarios: list[dict]) -> pd.DataFrame:
    """Maandtotalen per scenario, inclusief verschil t.o.v. de basisprognose."""
    months = sorted({d[:7] for d in table["Datum"]})
    value_cols = ["Verwacht (basis)"] + [sc["name"] for sc in scenarios]
    rows = []
    for ym in months:
        subset = table[table["Datum"].str.startswith(ym)]
        base = float(subset["Verwacht (basis)"].fillna(0).sum())
        for col in value_cols:
            total = float(subset[col].fillna(0).sum())
            rows.append({
                "Maand": ym,
                "Reeks": col,
                "Totaal": round(total, 2),
                "Verschil t.o.v. basis": round(total - base, 2),
                "Verschil %": round(100 * (total - base) / base, 2) if base else None,
            })
    return pd.DataFrame(rows)


def _meta_rows(forecast: dict, scenarios: list[dict]) -> list[tuple[str, str]]:
    m = forecast.get("metrics", {})
    return [
        ("Bestand", "KasVisie - kasprognose"),
        ("Model", forecast.get("modelLabel", "")),
        # Versie plus datavingerafdruk: hiermee is achteraf vast te stellen
        # welk model op welke dataset dit bestand heeft geproduceerd.
        ("Modelversie", forecast.get("modelVersion", "")),
        ("Datavingerafdruk", forecast.get("dataFingerprint", "")),
        ("Prognose vanaf", forecast.get("forecastStart", "")),
        ("Laatste realisatie", forecast.get("lastObservation", "")),
        ("Geëxporteerd op", date.today().isoformat()),
        ("Valuta", forecast.get("currency", "EUR")),
        ("Backtest: binnen band", f"{m.get('coverage_pct', '-')}% (doel {m.get('coverage_target', '-')}%)"),
        ("Backtest: pinball-loss", str(m.get("pinball", "-"))),
        ("Aantal scenario's", str(len(scenarios))),
        ("Let op", "Prognoses zijn schattingen; bijgestelde reeksen bevatten handmatige aannames."),
    ]


def to_csv(table: pd.DataFrame) -> bytes:
    """CSV voor Nederlandse Excel: puntkomma, decimale komma, BOM."""
    text = table.to_csv(index=False, sep=";", decimal=",", lineterminator="\r\n")
    return text.encode("utf-8-sig")


def _safe_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "-", name)[:31] or "Blad"


def to_xlsx(table: pd.DataFrame, forecast: dict, scenarios: list[dict]) -> bytes:
    buffer = io.BytesIO()
    totals = _totals(table, forecast, scenarios)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Prognose", index=False)
        totals.to_excel(writer, sheet_name="Totalen", index=False)

        meta = pd.DataFrame(_meta_rows(forecast, scenarios), columns=["Veld", "Waarde"])
        meta.to_excel(writer, sheet_name="Toelichting", index=False)

        money_cols = {c for c in table.columns if c not in
                      {"Datum", "Weekdag", "Soort", "Bijzonderheid"}}
        for sheet_name, frame in (("Prognose", table), ("Totalen", totals), ("Toelichting", meta)):
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            for idx, col in enumerate(frame.columns, start=1):
                letter = get_column_letter(idx)
                ws[f"{letter}1"].font = Font(bold=True)
                ws[f"{letter}1"].alignment = Alignment(vertical="top", wrap_text=True)
                width = max(len(str(col)) + 2, 14)
                ws.column_dimensions[letter].width = min(width, 42)
                if sheet_name == "Prognose" and col in money_cols:
                    for cell in ws[letter][1:]:
                        cell.number_format = '#,##0.00'
                if sheet_name == "Totalen" and col in {"Totaal", "Verschil t.o.v. basis"}:
                    for cell in ws[letter][1:]:
                        cell.number_format = '#,##0.00'
    return buffer.getvalue()


def filename(forecast: dict, ext: str) -> str:
    start = str(forecast.get("forecastStart", "")).replace("-", "")
    model = re.sub(r"[^a-z0-9]+", "-", str(forecast.get("model", "model")).lower())
    return f"kasvisie-prognose-{model}-{start}.{ext}"
